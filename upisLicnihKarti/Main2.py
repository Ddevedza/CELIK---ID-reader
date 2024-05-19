import ctypes
import os
import threading
import tkinter as tk
from tkinter import messagebox

import win32com.client
import win32com.client
from docxtpl import DocxTemplate
from openpyxl import load_workbook, Workbook

# Direktorijum skripte
script_dir = os.path.dirname(os.path.abspath(__file__))


# Kreira Excel po potrebi i updajtuje ga
def update_excel(data, folder_name, file_name):
    try:
        folder_path = os.path.join(script_dir, folder_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
            print(f"Folder created at: {folder_path}")

        excel_path = os.path.join(folder_path, file_name)
        wb = load_workbook(excel_path) if os.path.exists(excel_path) else Workbook()
        ws = wb.active

        if ws.max_row == 1:  # Daje heder ako nije napravljen
            headers = ['Serial Number', 'Name', 'ID Number', 'Authorized Validator']
            ws.append(headers)

        for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row, values_only=True):
            if data['jmbg'] == row[2]:
                print(f"ID {data['jmbg']} already exists. No update made.")
                return None, None  # Preskace duplirane unose

        next_row = ws.max_row + 1
        ws.append([next_row - 1, data['ime_prezime'], data['jmbg']])
        wb.save(excel_path)
        print(f"Workbook saved at: {excel_path}")
        return next_row - 1, folder_path

    except Exception as e:
        messagebox.showerror("Error in Excel Update", str(e))
        return None, None


def get_card_reader_driver():
    wmi = win32com.client.GetObject("winmgmts:")
    for item in wmi.InstancesOf("Win32_PnPEntity"):
        if "Card Reader" in item.Name or "card reader" in item.Name:
            print(item.Name)
            break


card_Name = get_card_reader_driver()


def id_card():
    while (True):
        try:
            # Učitavanje Čelik API DLL
            celik_api = ctypes.cdll.LoadLibrary(os.getcwd() + "\\CelikApi.dll")
            # Preuzimanje konstante
            EID_MAX_DocRegNo = 9
            EID_MAX_DocumentType = 2
            EID_MAX_IssuingDate = 10
            EID_MAX_ExpiryDate = 10
            EID_MAX_IssuingAuthority = 100
            EID_MAX_DocumentSerialNumber = 10
            EID_MAX_ChipSerialNumber = 14
            EID_MAX_DocumentName = 100
            EID_MAX_PersonalNumber = 13
            EID_MAX_Surname = 200
            EID_MAX_GivenName = 200
            EID_MAX_ParentGivenName = 200
            EID_MAX_Sex = 2
            EID_MAX_PlaceOfBirth = 200
            EID_MAX_StateOfBirth = 200
            EID_MAX_DateOfBirth = 10
            EID_MAX_CommunityOfBirth = 200
            EID_MAX_NationalityFull = 200

            EID_MAX_State = 100
            EID_MAX_Community = 200
            EID_MAX_Place = 200
            EID_MAX_Street = 200
            EID_MAX_HouseNumber = 20
            EID_MAX_HouseLetter = 8
            EID_MAX_Entrance = 10
            EID_MAX_Floor = 6
            EID_MAX_ApartmentNumber = 12
            EID_MAX_AddressDate = 10
            EID_MAX_AddressLabel = 60

            # Definisanje struktura
            class EID_DOCUMENT_DATA(ctypes.Structure):
                _fields_ = [
                    ("docRegNo", ctypes.c_char * EID_MAX_DocRegNo),
                    ("docRegNoSize", ctypes.c_int),
                    ("documentType", ctypes.c_char * EID_MAX_DocumentType),
                    ("documentTypeSize", ctypes.c_int),
                    ("issuingDate", ctypes.c_char * EID_MAX_IssuingDate),
                    ("issuingDateSize", ctypes.c_int),
                    ("expiryDate", ctypes.c_char * EID_MAX_ExpiryDate),
                    ("expiryDateSize", ctypes.c_int),
                    ("issuingAuthority", ctypes.c_char * EID_MAX_IssuingAuthority),
                    ("issuingAuthoritySize", ctypes.c_int),
                    ("documentSerialNumber", ctypes.c_char * EID_MAX_DocumentSerialNumber),
                    ("documentSerialNumberSize", ctypes.c_int),
                    ("chipSerialNumber", ctypes.c_char * EID_MAX_ChipSerialNumber),
                    ("chipSerialNumberSize", ctypes.c_int),
                    ("documentName", ctypes.c_char * EID_MAX_DocumentName),
                    ("documentNameSize", ctypes.c_int)
                ]

            class EID_FIXED_PERSONAL_DATA(ctypes.Structure):
                _fields_ = [
                    ("personalNumber", ctypes.c_char * EID_MAX_PersonalNumber),
                    ("personalNumberSize", ctypes.c_int),
                    ("surname", ctypes.c_char * EID_MAX_Surname),
                    ("surnameSize", ctypes.c_int),
                    ("givenName", ctypes.c_char * EID_MAX_GivenName),
                    ("givenNameSize", ctypes.c_int),
                    ("parentGivenName", ctypes.c_char * EID_MAX_ParentGivenName),
                    ("parentGivenNameSize", ctypes.c_int),
                    ("sex", ctypes.c_char * EID_MAX_Sex),
                    ("sexSize", ctypes.c_int),
                    ("placeOfBirth", ctypes.c_char * EID_MAX_PlaceOfBirth),
                    ("placeOfBirthSize", ctypes.c_int),
                    ("stateOfBirth", ctypes.c_char * EID_MAX_StateOfBirth),
                    ("stateOfBirthSize", ctypes.c_int),
                    ("dateOfBirth", ctypes.c_char * EID_MAX_DateOfBirth),
                    ("dateOfBirthSize", ctypes.c_int),
                    ("communityOfBirth", ctypes.c_char * EID_MAX_CommunityOfBirth),
                    ("communityOfBirthSize", ctypes.c_int),
                    ("nationalityFull", ctypes.c_char * EID_MAX_NationalityFull),
                    ("nationalityFullSize", ctypes.c_int)
                ]

            class EID_VARIABLE_PERSONAL_DATA(ctypes.Structure):
                _fields_ = [
                    ("state", ctypes.c_char * EID_MAX_State),
                    ("stateSize", ctypes.c_int),
                    ("community", ctypes.c_char * EID_MAX_Community),
                    ("communitySize", ctypes.c_int),
                    ("place", ctypes.c_char * EID_MAX_Place),
                    ("placeSize", ctypes.c_int),
                    ("street", ctypes.c_char * EID_MAX_Street),
                    ("streetSize", ctypes.c_int),
                    ("houseNumber", ctypes.c_char * EID_MAX_HouseNumber),
                    ("houseNumberSize", ctypes.c_int),
                    ("houseLetter", ctypes.c_char * EID_MAX_HouseLetter),
                    ("houseLetterSize", ctypes.c_int),
                    ("entrance", ctypes.c_char * EID_MAX_Entrance),
                    ("entranceSize", ctypes.c_int),
                    ("floor", ctypes.c_char * EID_MAX_Floor),
                    ("floorSize", ctypes.c_int),
                    ("apartmentNumber", ctypes.c_char * EID_MAX_ApartmentNumber),
                    ("apartmentNumberSize", ctypes.c_int),
                    ("addressDate", ctypes.c_char * EID_MAX_AddressDate),
                    ("addressDateSize", ctypes.c_int),
                    ("addressLabel", ctypes.c_char * EID_MAX_AddressLabel),
                    ("addressLabelSize", ctypes.c_int)
                ]

            celik_api.EidSetOption.argtypes = [ctypes.c_int, ctypes.c_uint]
            celik_api.EidSetOption.restype = ctypes.c_int

            EidSetOption_arg1 = 1
            EidSetOption_arg2 = 0

            EidSetOption_function_call_result = celik_api.EidSetOption(EidSetOption_arg1, EidSetOption_arg2)
            if EidSetOption_function_call_result != 0:
                print("EidSetOption_function_call_result runtime exception")

            # Definisanje prototipa funkcija
            celik_api.EidStartup.argtypes = [ctypes.c_int]
            celik_api.EidStartup.restype = ctypes.c_int

            celik_api.EidBeginRead.argtypes = [ctypes.c_char_p, ctypes.POINTER(ctypes.c_int)]
            celik_api.EidBeginRead.restype = ctypes.c_int

            celik_api.EidReadDocumentData.argtypes = [ctypes.POINTER(EID_DOCUMENT_DATA)]
            celik_api.EidReadDocumentData.restype = ctypes.c_int

            celik_api.EidReadFixedPersonalData.argtypes = [ctypes.POINTER(EID_FIXED_PERSONAL_DATA)]
            celik_api.EidReadFixedPersonalData.restype = ctypes.c_int

            celik_api.EidReadVariablePersonalData.argtypes = [ctypes.POINTER(EID_VARIABLE_PERSONAL_DATA)]
            celik_api.EidReadVariablePersonalData.restype = ctypes.c_int

            celik_api.EidEndRead.restype = ctypes.c_int
            celik_api.EidCleanup.restype = ctypes.c_int

            # Pokrenite API
            startup_result = celik_api.EidStartup(4)  # Verzija API-ja je 4
            if startup_result != 0:
                print(f"Neuspešno pokretanje API-ja, greška {startup_result}")
                exit()

            card_type = ctypes.c_int(0)

            # definise ime licne karte na osnovu drajvera (Odgovarajuci drajver mora biti instaliran)
            ediBeginRead_result = celik_api.EidBeginRead(card_Name, ctypes.byref(card_type))
            print("ediBeginRead_result " + str(ediBeginRead_result))

            if ediBeginRead_result == 0:
                document_data = EID_DOCUMENT_DATA()
                fixed_data = EID_FIXED_PERSONAL_DATA()
                variable_data = EID_VARIABLE_PERSONAL_DATA()

                result_document_data = celik_api.EidReadDocumentData(ctypes.byref(document_data))
                result_fixed_data = celik_api.EidReadFixedPersonalData(ctypes.byref(fixed_data))
                result_variable_data = celik_api.EidReadVariablePersonalData(ctypes.byref(variable_data))

                if result_document_data == 0 and result_fixed_data == 0 and result_variable_data == 0:
                    desktop_dir = os.path.join(os.environ['USERPROFILE'], 'Desktop')

                    ime_prezime = fixed_data.givenName.decode('utf-8') + " " + fixed_data.surname.decode('utf-8')
                    adresa = f"{variable_data.street.decode('utf-8')} {variable_data.houseNumber.decode('utf-8')}"

                    print(ime_prezime)
                    context = {
                        'ime_prezime': ime_prezime,
                        'br1': fixed_data.personalNumber.decode('utf-8')[0],
                        'br2': fixed_data.personalNumber.decode('utf-8')[1],
                        'br3': fixed_data.personalNumber.decode('utf-8')[2],
                        'br4': fixed_data.personalNumber.decode('utf-8')[3],
                        'br5': fixed_data.personalNumber.decode('utf-8')[4],
                        'br6': fixed_data.personalNumber.decode('utf-8')[5],
                        'br7': fixed_data.personalNumber.decode('utf-8')[6],
                        'br8': fixed_data.personalNumber.decode('utf-8')[7],
                        'br9': fixed_data.personalNumber.decode('utf-8')[8],
                        'br10': fixed_data.personalNumber.decode('utf-8')[9],
                        'br11': fixed_data.personalNumber.decode('utf-8')[10],
                        'br12': fixed_data.personalNumber.decode('utf-8')[11],
                        'br13': fixed_data.personalNumber.decode('utf-8')[12],
                        'dan_rodjenja': fixed_data.dateOfBirth.decode('utf-8'),
                        'adresa_prebivalista': adresa,
                        'mesto': variable_data.place.decode('utf-8')

                    }

                    # Sta upisujemo u excel
                    data = {
                        'ime_prezime': ime_prezime,
                        'jmbg': fixed_data.personalNumber.decode('utf-8'),
                    }
                    print('ad')
                    while celik_api.EidCleanup() != 0:
                        while celik_api.EidEndRead() != 0:
                            return data, context





            else:
                print("Neuspešno započinjanje čitanja sa čitačem")
                messagebox.showinfo("Running",
                                    "Neuspešno započinjanje čitanja sa čitačem, proverite da li ste lepo uneli karticu!")

        except Exception as e:
            messagebox.showerror("Greska!", str(e))


def run_script():
    try:
        data, context = id_card()
        if not data:
            return

        redni_broj, folder_path = update_excel(data, "Data Folder", "data.xlsx")
        if not redni_broj:
            return

        doc = DocxTemplate(os.path.join(script_dir, "template.docx"))
        doc.render(context)
        doc.save(os.path.join(folder_path, f"{redni_broj}_document.docx"))
        messagebox.showinfo("Uspeh", "Podaci uspesno sacuvani.")
    except Exception as e:
        messagebox.showerror("Greska u skripti", str(e))


def start_thread():
    thread = threading.Thread(target=run_script)
    thread.start()


app = tk.Tk()
app.title("Celik-ID READER")

frame = tk.Frame(app)
frame.pack(padx=10, pady=10)

btn_start = tk.Button(frame, text="Istampaj podatke", command=start_thread)
btn_start.pack(fill=tk.BOTH, expand=True)

app.mainloop()
